import io
import json
import os

import requests
from openpyxl import load_workbook
import re
import psycopg2
# from PyPDF2 import PdfReader  # 用于读取PDF文件
import fitz
import subprocess
import os

def clean_file_name(name):
    return re.sub(r'[<>:"/\\|?*]', '', name)


def download_pdf(save_path, pdf_name, pdf_url):
    send_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36",
        "Connection": "keep-alive",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.8"}

    response = requests.get(pdf_url, headers=send_headers)
    bytes_io = io.BytesIO(response.content)
    cleaned_pdf_name = clean_file_name(pdf_name)
    pdf_file_path = os.path.join(save_path, f"{cleaned_pdf_name}.PDF")

    with open(pdf_file_path, mode='wb') as f:
        f.write(bytes_io.getvalue())
        print(f'{cleaned_pdf_name}.PDF,下载成功！')

    return pdf_file_path  # 返回文件路径


import subprocess
import os
import requests


def translate_pdf_content(pdf_path):
    # 定义输出路径
    global translated_part
    output_dir = 'D:/shix_2024/paper_web/static/papers'#经过pdf_kit库处理后的文件存放路径
    magic_pdf_command = f'magic-pdf -p "{pdf_path}" -o "{output_dir}"'

    try:
        subprocess.run(magic_pdf_command, shell=True, check=True)

        # 根据 PDF 文件名构建 .md 文件的路径
        pdf_name = os.path.basename(pdf_path).replace('.PDF', '')  # 获取 PDF 文件名并去掉扩展名
        md_file_path = os.path.join(output_dir, pdf_name, 'auto', f'{pdf_name}.md')  # 拼接出 .md 文件的完整路径

        if os.path.exists(md_file_path):
            with open(md_file_path, 'r', encoding='utf-8') as md_file:
                text = ""
                lines = md_file.readlines()  # 将文件的所有行读取到一个列表中
            for i in range(len(lines)):
                line = lines[i]

                # 检查当前行是否包含图片语句
                if line.strip().startswith("![]("):
                    # 提取图片的路径
                    start_index = line.find("(") + 1
                    end_index = line.find(")")
                    if start_index > 0 and end_index > start_index:
                        img_path = line[start_index:end_index]
                        # 将图片语句转换为 HTML <img> 标签
                        img_html = f'<img src="../static/papers/{pdf_name}/auto/{img_path}" alt="Image" />\n'
                        text += img_html  # 将 HTML 标签添加到文本中
                elif lines[i - 1].strip() == "$$" and lines[i + 1].strip() == "$$":  # 确保不越界
                    continue  # 如果前后行都是"$"，则跳过当前行
                elif line.strip() == "$$":  # 检查当前行是否为"$"
                    # 检查当前行后面的两行
                    if i < len(lines) - 2:  # 确保不越界
                        second_line = lines[i + 1].strip()  # 第二行
                        third_line = lines[i + 2].strip()  # 第三行
                        # 检查第二行是否不为空，并确保第三行是"$"
                        if second_line and third_line == "$$":  # 如果第二行不为空
                            # 将前一行、第二行和第三行合并为一行输出
                            combined_line = f"{lines[i].strip()} {second_line} {third_line}\n"
                            text += combined_line
                else:
                    text += line

            # 根据换行符将文本分割成小部分
            parts = text.splitlines()  # 按行分割
            translated_parts = []

            # Dify翻译API信息
            api_url = "http://localhost/v1/workflows/run"
            api_key = "app-HjTLvDPziSu8OAvh0umxsOSF"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }

            # 循环每一部分进行翻译
            for part in parts:
                if part.strip():  # 确保不处理空行
                    data = {
                        "inputs": {'yuan': part},
                        "user": "abc-123"
                    }
                    try:
                        response = requests.post(api_url, headers=headers, data=json.dumps(data))
                        response.raise_for_status()  # 检查请求是否成功

                        outputs = response.json().get("data", {}).get("outputs", {})
                        translated_part0 = outputs.get("yuan")
                        translated_part1 = outputs.get("text")

                        if translated_part0 and translated_part1:  # 当两个都有值时
                            translated_part = f"{translated_part0}{translated_part1}\n"
                        elif translated_part0:  # 只有 yuan 有值
                            translated_part = f"{translated_part0}\n"
                        elif translated_part1:  # 只有 text 有值
                            translated_part = f"{translated_part1}\n"
                        else:  # 两个都没有值
                            translated_part = "没有返回值"
                    except Exception as e:
                        print(f"翻译过程中发生错误: {e}")
                        translated_part = "翻译失败"

                    translated_parts.append(translated_part)  # 将翻译结果添加到列表中

            # 合并所有翻译部分，保持换行符
            translated_text = '\n'.join(translated_parts)

        else:
            print(f"未找到文件: {md_file_path}")
            translated_text = ""

    except subprocess.CalledProcessError as e:
        print(f"An error occurred while extracting text: {e}")
        translated_text = ""

    return translated_text


def store_url_in_db(pdf_name, author, abstract, pdf_url, transformed):
    try:
        connection = psycopg2.connect(
            dbname='paper_db',
            user='postgres',
            password='difyai123456',
            host='localhost',
            port='5432'
        )
        cursor = connection.cursor()

        # 插入PDF信息和翻译内容到数据库
        insert_query = """INSERT INTO paper_data (pdf_name, author, abstract, pdf_url, transformed) VALUES (%s, %s, %s, %s, %s)"""
        cursor.execute(insert_query, (pdf_name, author, abstract, pdf_url, transformed))

        connection.commit()
        print(f'URL {pdf_url} 存储成功！翻译内容也已存储。')

    except Exception as e:
        print(f'错误: {e}')

    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


def store_url_in_db1(translated,title):
    try:
        connection = psycopg2.connect(
            dbname='paper_db',
            user='postgres',
            password='difyai123456',
            host='localhost',
            port='5432'
        )
        cursor = connection.cursor()


        # 使用标题查找指定的行，并更新翻译内容
        update_query = """UPDATE paper_pdf SET translated = %s WHERE translated_title = %s"""
        cursor.execute(update_query, (translated, title))

        connection.commit()
        print(f'存储成功！翻译内容已更新。')

    except Exception as e:
        print(f'错误: {e}')

    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

def main():
    save_path = 'D:/shix_2024/paper_PC/sxdmx_pdf/'
    excel_file = 'D:/shix_2024/paper_PC/时序大模型.xlsx'
    max_loop_count = 43

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    loop_count = 0
    for row in sheet.iter_rows(min_row=2):
        pdf_url = row[3].value
        pdf_name = row[1].value

        if pdf_url:
            if pdf_name:
                pdf_name = str(pdf_name)
                pdf_path = download_pdf(save_path, pdf_name, pdf_url)
                translated_text = translate_pdf_content(pdf_path)  # 读取并翻译PDF内容
                store_url_in_db(pdf_name, pdf_url, translated_text)  # 存储URL和翻译内容到数据库
                loop_count += 1
            else:
                print("第%s行文件名为空，跳过..." % (row[0].row))

        if loop_count >= max_loop_count:
            break


if __name__ == '__main__':
    main()
